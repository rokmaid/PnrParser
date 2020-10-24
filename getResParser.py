import xml.dom.minidom 
import pandas as pd 
from openpyxl import workbook
from openpyxl.workbook.workbook import Workbook


class Counter:
    _val=1  
    def increment(self):
      self._val+=1
 
    def getVal(self):
     return self._val 


def parsePNR(resultText):


 
 counter = Counter()
 
 book = Workbook()
 sheet = book.active

 #xmlfile = open("C:/Users/SG0208525/Python Course/PNRParser/GetReservationRS.xml", "r")
 
 dom=xml.dom.minidom.parseString(resultText)

 if getErrors(dom)!=True:
    
   segments =dom.getElementsByTagName("stl19:Segment")
   pq_list=dom.getElementsByTagName("PriceQuoteInfo")
   passengers=dom.getElementsByTagName("stl19:Passenger")
 
 #print("Segments ID : {0}".format( dom.getElementsByTagName("stl19:Segment")[0].getAttribute("sequence") )) 
   pnr = dom.getElementsByTagName("stl19:RecordLocator")[0].firstChild.nodeValue
   agent_sine= dom.getElementsByTagName("stl19:CreationAgentID")[0].firstChild.nodeValue
   print("PNR : {0}".format(pnr))
   print("Agent Sine : {0}".format(agent_sine))
 
 #Pnr Header 
   sheet.cell(row=1, column=counter.getVal()).value = "PNR"
 #Pnr value 
   sheet.cell(row=2, column=counter.getVal()).value = pnr
   counter.increment()
 #Agent sine Header
   sheet.cell(row=1, column=counter.getVal()).value = "Agent Sine"
 #Agent sine value 
   sheet.cell(row=2, column=counter.getVal()).value = agent_sine
   counter.increment()

   getSegments(segments,sheet,counter)
   getPQinfo(pq_list,sheet,counter)

   getPassengers(passengers,sheet,counter)


 # save excel file 

   book.save(pnr+".xlsx")

def getSegments(segment_list,workbook,counter):
   print("Segment Info ")
   
   for seg in segment_list:
       
       id=seg.getAttribute("sequence")
       #ID Header
       workbook.cell(row=1,column=counter.getVal()).value="Segment Number"
       #ID value 
       workbook.cell(row=2,column=counter.getVal()).value=id
       counter.increment()
       departure=seg.getElementsByTagName("stl19:DepartureAirport")[0].firstChild.nodeValue
       #Depature Header 
       workbook.cell(row=1,column=counter.getVal()).value="Departure"
       #Departure value 
       workbook.cell(row=2,column=counter.getVal()).value=departure
       counter.increment()
       arrival = seg.getElementsByTagName("stl19:ArrivalAirport")[0].firstChild.nodeValue 
        #Arrival Header 
       workbook.cell(row=1,column=counter.getVal()).value="Arrival"
       #Arrival Value 
       workbook.cell(row=2,column=counter.getVal()).value=arrival
       counter.increment()

       flightnumber=seg.getElementsByTagName("stl19:FlightNumber")[0].firstChild.nodeValue
       #Flight number header 
       workbook.cell(row=1,column=counter.getVal()).value="Flight Number"
        #Flight number value  
       workbook.cell(row=2,column=counter.getVal()).value=flightnumber
       counter.increment()

       print("ID : {0}".format(id))
       print("Departure : {0}".format(departure))
       print("Departure : {0}".format(arrival)) 

       print("Column Count : "+ str(counter.getVal()))

def getPassengers(pax_list,workbook,counter):
    print("Pax s  ")
    
    rowcounter = 1 
    for pax in pax_list:
        nameref=pax.getAttribute("nameId") 
        name=pax.getElementsByTagName("stl19:LastName")[0].firstChild.nodeValue+"/"+pax.getElementsByTagName("stl19:FirstName")[0].firstChild.nodeValue
        workbook.cell(row=1,column=counter.getVal()).value="name Ref" 
        workbook.cell(row=2,column=counter.getVal()).value=nameref
        counter.increment()

        workbook.cell(row=1,column=counter.getVal()).value="Name"
        workbook.cell(row=2,column=counter.getVal()).value=name
        counter.increment()

        print("Column Count : "+ str(counter.getVal()))

def getPQinfo(pq_list,workbook,counter):
   print("PQ info ")
   for pq in pq_list:
       index = pq.getElementsByTagName("PriceQuote")[0].getAttribute("number")
       currency= pq.getElementsByTagName("Amounts")[0].getElementsByTagName("Total")[0].getAttribute("currencyCode")
       farebasis=pq.getElementsByTagName("SegmentInfo")[0].getElementsByTagName("FareBasis")[0].firstChild.nodeValue
       basefare=pq.getElementsByTagName("FareInfo")[0].getElementsByTagName("BaseFare")[0].firstChild.nodeValue
       totaltax=pq.getElementsByTagName("FareInfo")[0].getElementsByTagName("TotalTax")[0].firstChild.nodeValue
       total=pq.getElementsByTagName("FareInfo")[0].getElementsByTagName("TotalFare")[0].firstChild.nodeValue
       workbook.cell(row=1,column=counter.getVal()).value="PQ Index" 
       workbook.cell(row=2,column=counter.getVal()).value=index 
       counter.increment()
       print("Index : {0}".format(index))

       print("Currency : {0}".format(currency))
       workbook.cell(row=1,column=counter.getVal()).value="Currency" 
       workbook.cell(row=2,column=counter.getVal()).value=currency
       counter.increment()

       print("Fare Basis Code : {0}".format(farebasis))
       workbook.cell(row=1,column=counter.getVal()).value="Fare Basis" 
       workbook.cell(row=2,column=counter.getVal()).value=farebasis
       counter.increment()
       print("Base Fare : {0}".format(basefare))
       workbook.cell(row=1,column=counter.getVal()).value="Base Fare" 
       workbook.cell(row=2,column=counter.getVal()).value=basefare
       counter.increment()
       print("Total Tax : {0}".format(totaltax))
       workbook.cell(row=1,column=counter.getVal()).value="Total Tax" 
       workbook.cell(row=2,column=counter.getVal()).value=totaltax
       counter.increment()
       print("Total : {0}".format(total))
       workbook.cell(row=1,column=counter.getVal()).value="Total" 
       workbook.cell(row=2,column=counter.getVal()).value=total
       counter.increment()

def getErrors(doc):
  
  #Look for USG Errors or no PNR in AAA before parsing the PNR and creating the excel file 

  faullist= doc.getElementsByTagName("soap-env:Fault")
  errors=doc.getElementsByTagName("stl19:Errors")  
  if len(faullist)>0:
     faultstring=faullist[0].getElementsByTagName("faultstring")[0].firstChild.nodeValue

     print("### Error #### {0}".format(faultstring)) 
     return True
  elif len(errors) > 0 :
     error = errors[0].getElementsByTagName("stl19:Error")[0] 
     message =error.getElementsByTagName("stl19:Message")[0].firstChild.nodeValue 
     print("#### Error #### {0}".format(message))
     return True      
  # else if to check if there is no PNR in AAA 
  return False

#parsePNR(xml);    